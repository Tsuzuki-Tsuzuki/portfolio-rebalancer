"""
Portfolio Rebalancing Script
Goal: Adjust weights of 12-stock portfolio to achieve target beta = 1.2
Constraints:
  - Weights sum to 100%
  - No short selling (all weights >= 0)
  - Max weight per stock = 20%
  - Minimize deviation from current weights
"""

import numpy as np
from scipy.optimize import minimize
import pandas as pd

# ─────────────────────────────────────────────
# 1. DATA — Portfolio stocks, betas, current weights
# ─────────────────────────────────────────────

stocks = ['AKDITAL', 'AWB', 'COSUMAR', 'TGCC', 'ALLIANCES', 'TAQA',
          'ADDOHA', 'LABELVIE', 'LAFARGEHO', 'MARSA', 'IAM', 'JET']

betas = np.array([
    1.898229001,   # AKDITAL
    0.967391147,   # AWB (Attijariwafa Bank)
    1.233379807,   # COSUMAR
    1.204639271,   # TGCC
    1.196411641,   # ALLIANCES
    0.928444943,   # TAQA MOROCCO
    1.226712211,   # ADDOHA
    1.507428458,   # LABEL VIE
    2.679425695,   # LAFARGEHO
    1.638600079,   # MARSA MAROC
    1.123400677,   # IAM
    0.814617961,   # JET CONTRACTORS
])

current_weights = np.array([
    0.0667,   # AKDITAL
    0.1043,   # AWB
    0.0474,   # COSUMAR
    0.1203,   # TGCC
    0.0465,   # ALLIANCES
    0.1945,   # TAQA MOROCCO
    0.0626,   # ADDOHA
    0.0655,   # LABEL VIE
    0.0931,   # LAFARGEHO
    0.0899,   # MARSA MAROC
    0.0635,   # IAM
    0.0457,   # JET CONTRACTORS
])

TARGET_BETA = 1.2
MAX_WEIGHT  = 0.20   # 20% cap per stock
INVESTMENT  = 1_000_000   # MAD

# ─────────────────────────────────────────────
# 2. OPTIMIZATION
# Objective: minimize sum of squared changes from current weights
# ─────────────────────────────────────────────

def objective(w):
    return np.sum((w - current_weights) ** 2)

def objective_grad(w):
    return 2 * (w - current_weights)

constraints = [
    # Portfolio beta must equal target
    {
        'type': 'eq',
        'fun': lambda w: np.dot(w, betas) - TARGET_BETA
    },
    # Weights must sum to 1
    {
        'type': 'eq',
        'fun': lambda w: np.sum(w) - 1.0
    }
]

bounds = [(0.0, MAX_WEIGHT)] * len(stocks)   # No short selling, max 20%

result = minimize(
    objective,
    x0=current_weights,          # Start from current weights
    method='SLSQP',
    jac=objective_grad,
    bounds=bounds,
    constraints=constraints,
    options={'ftol': 1e-10, 'maxiter': 1000}
)

# ─────────────────────────────────────────────
# 3. RESULTS
# ─────────────────────────────────────────────

if not result.success:
    print(f"⚠️  Optimization warning: {result.message}")

new_weights = result.x

old_beta = np.dot(current_weights, betas)
new_beta = np.dot(new_weights, betas)

print("=" * 65)
print("         PORTFOLIO REBALANCING RESULTS")
print("=" * 65)
print(f"  Current Portfolio Beta : {old_beta:.4f}")
print(f"  Target  Portfolio Beta : {TARGET_BETA:.4f}")
print(f"  Achieved Portfolio Beta: {new_beta:.4f}")
print(f"  Sum of new weights     : {np.sum(new_weights):.6f}")
print("=" * 65)

print(f"\n{'Stock':<15} {'β':>8} {'Old Weight':>12} {'New Weight':>12} {'Change':>10} {'New Capital (MAD)':>20}")
print("-" * 80)

for i, stock in enumerate(stocks):
    delta = new_weights[i] - current_weights[i]
    capital = new_weights[i] * INVESTMENT
    arrow = "▲" if delta > 0.0001 else ("▼" if delta < -0.0001 else "─")
    print(f"{stock:<15} {betas[i]:>8.4f} {current_weights[i]:>11.2%} {new_weights[i]:>11.2%} {arrow} {delta:>+8.2%} {capital:>18,.0f}")

print("-" * 80)
print(f"{'TOTAL':<15} {'':>8} {np.sum(current_weights):>11.2%} {np.sum(new_weights):>11.2%} {'':>10} {np.sum(new_weights)*INVESTMENT:>18,.0f}")

# ─────────────────────────────────────────────
# 4. EXPORT TO EXCEL
# ─────────────────────────────────────────────

df = pd.DataFrame({
    'Stock': stocks,
    'Beta (β)': betas,
    'Current Weight': current_weights,
    'New Weight (Rebalanced)': new_weights,
    'Weight Change': new_weights - current_weights,
    'Current Capital (MAD)': current_weights * INVESTMENT,
    'New Capital (MAD)': new_weights * INVESTMENT,
    'Capital Change (MAD)': (new_weights - current_weights) * INVESTMENT,
})

# Round for cleanliness
df['Current Weight'] = df['Current Weight'].map(lambda x: f"{x:.2%}")
df['New Weight (Rebalanced)'] = df['New Weight (Rebalanced)'].map(lambda x: f"{x:.2%}")
df['Weight Change'] = df['Weight Change'].map(lambda x: f"{x:+.2%}")
df['Current Capital (MAD)'] = df['Current Capital (MAD)'].map(lambda x: f"{x:,.0f}")
df['New Capital (MAD)'] = df['New Capital (MAD)'].map(lambda x: f"{x:,.0f}")
df['Capital Change (MAD)'] = df['Capital Change (MAD)'].map(lambda x: f"{x:+,.0f}")

# Summary row
summary = pd.DataFrame([{
    'Stock': 'PORTFOLIO TOTAL',
    'Beta (β)': f"{new_beta:.4f}",
    'Current Weight': f"{np.sum(current_weights):.2%}",
    'New Weight (Rebalanced)': f"{np.sum(new_weights):.2%}",
    'Weight Change': '—',
    'Current Capital (MAD)': f"{INVESTMENT:,.0f}",
    'New Capital (MAD)': f"{np.sum(new_weights)*INVESTMENT:,.0f}",
    'Capital Change (MAD)': '—',
}])

df_out = pd.concat([df, summary], ignore_index=True)

output_path = "rebalanced_portfolio_beta12.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_out.to_excel(writer, index=False, sheet_name='Rebalancing Results')

    ws = writer.sheets['Rebalancing Results']

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    col_widths = [18, 10, 16, 22, 14, 22, 20, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Row heights
    ws.row_dimensions[1].height = 30

    # Total row styling
    total_row = ws.max_row
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    for cell in ws[total_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True, name="Arial", size=10)

    # Zebra striping + border
    thin = Side(style='thin', color='CCCCCC')
    border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        if row[0].row % 2 == 0:
            for cell in row:
                if row[0].row != total_row:
                    cell.fill = PatternFill("solid", fgColor="F2F7FC")

    # Summary info above table
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws['A1'] = f"Portfolio Rebalancing — Target Beta: {TARGET_BETA}"
    ws['A1'].font = Font(bold=True, name="Arial", size=12, color="1F3864")
    ws['A2'] = f"Current Beta: {old_beta:.4f}   →   Achieved Beta: {new_beta:.4f}   |   Total Investment: 1,000,000 MAD"
    ws['A2'].font = Font(italic=True, name="Arial", size=10, color="444444")

print(f"\n✅ Excel file exported to: {output_path}")
