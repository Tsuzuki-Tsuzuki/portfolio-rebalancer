"""
Portfolio Rebalancing Web App
Target: achieve a portfolio beta of 1.2 by adjusting stock weights
"""

import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from scipy.optimize import minimize
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Portfolio Beta Rebalancer",
    page_icon="📊",
    layout="wide",
)

st.title("Portfolio Beta Rebalancer")
st.markdown(
    "Adjust your portfolio weights to reach a **target beta of 1.2** "
    "while keeping weight changes as small as possible."
)

# ─────────────────────────────────────────────
# Sidebar — global settings
# ─────────────────────────────────────────────

with st.sidebar:
    st.header("Settings")
    target_beta = st.number_input("Target Beta", value=1.20, min_value=0.01, step=0.01, format="%.2f")
    max_weight  = st.slider("Max weight per stock (%)", min_value=5, max_value=100, value=20)
    investment  = st.number_input("Total investment (MAD)", value=1_000_000, step=100_000, min_value=0)
    st.markdown("---")
    st.markdown(
        "**How it works**\n\n"
        "The optimizer minimizes the total squared change in weights "
        "subject to:\n"
        "- Weights sum to 100 %\n"
        "- No short selling (weight ≥ 0)\n"
        "- Per-stock weight cap\n"
        "- Portfolio beta = target\n"
    )

# ─────────────────────────────────────────────
# Input mode tabs
# ─────────────────────────────────────────────

tab_upload, tab_manual = st.tabs(["Upload Excel file", "Enter data manually"])

df_input = None  # will hold a DataFrame with columns: Stock, Beta, Weight

# ── Tab 1: Excel upload ──────────────────────

with tab_upload:
    st.subheader("Upload your portfolio")
    st.markdown(
        "Your Excel file must contain **at least three columns**: "
        "`Stock` (name), `Beta`, and `Weight` (as a decimal, e.g. 0.15 for 15 %, "
        "or as a percentage, e.g. 15).\n\n"
        "The column headers are **not case-sensitive** and extra columns are ignored."
    )

    template_buf = io.BytesIO()
    template_df = pd.DataFrame({
        "Stock":  ["AKDITAL", "AWB", "COSUMAR", "TGCC"],
        "Beta":   [1.8982,    0.9674, 1.2334,   1.2046],
        "Weight": [0.0667,    0.1043, 0.0474,   0.1203],
    })
    with pd.ExcelWriter(template_buf, engine="openpyxl") as w:
        template_df.to_excel(w, index=False, sheet_name="Portfolio")
    template_buf.seek(0)
    st.download_button(
        "Download template (.xlsx)",
        data=template_buf,
        file_name="portfolio_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    uploaded = st.file_uploader("Choose an Excel file (.xlsx)", type=["xlsx", "xls"])

    if uploaded:
        try:
            raw = pd.read_excel(uploaded)
            raw.columns = [c.strip().lower() for c in raw.columns]

            # Flexible column detection
            col_map = {}
            for col in raw.columns:
                if "stock" in col or "name" in col or "ticker" in col or "symbol" in col:
                    col_map["Stock"] = col
                elif "beta" in col:
                    col_map["Beta"] = col
                elif "weight" in col or "alloc" in col or "share" in col:
                    col_map["Weight"] = col

            missing = [k for k in ("Stock", "Beta", "Weight") if k not in col_map]
            if missing:
                st.error(f"Could not find columns for: {', '.join(missing)}. "
                         f"Found columns: {list(raw.columns)}")
            else:
                df_input = raw[[col_map["Stock"], col_map["Beta"], col_map["Weight"]]].copy()
                df_input.columns = ["Stock", "Beta", "Weight"]
                df_input = df_input.dropna()
                df_input["Beta"]   = pd.to_numeric(df_input["Beta"],   errors="coerce")
                df_input["Weight"] = pd.to_numeric(df_input["Weight"], errors="coerce")
                df_input = df_input.dropna()

                # Normalise weights: if user entered percentages (sum ~ 100) convert to decimals
                if df_input["Weight"].sum() > 1.5:
                    df_input["Weight"] = df_input["Weight"] / 100.0

                st.success(f"Loaded {len(df_input)} stocks.")
                st.dataframe(df_input.style.format({"Beta": "{:.4f}", "Weight": "{:.2%}"}),
                             use_container_width=True)
        except Exception as e:
            st.error(f"Error reading file: {e}")

# ── Tab 2: Manual entry ──────────────────────

with tab_manual:
    st.subheader("Enter your portfolio manually")
    st.markdown(
        "Fill in your stocks below. Enter **Beta** as a number (e.g. 1.25) "
        "and **Weight** as a percentage (e.g. 15 for 15 %)."
    )

    # Default data so there's something to start with
    default_data = {
        "Stock":  ["AKDITAL", "AWB", "COSUMAR", "TGCC", "ALLIANCES",
                   "TAQA", "ADDOHA", "LABELVIE", "LAFARGEHO", "MARSA", "IAM", "JET"],
        "Beta":   [1.898229, 0.967391, 1.233380, 1.204639, 1.196412,
                   0.928445, 1.226712, 1.507428, 2.679426, 1.638600, 1.123401, 0.814618],
        "Weight (%)": [6.67, 10.43, 4.74, 12.03, 4.65,
                       19.45, 6.26, 6.55, 9.31, 8.99, 6.35, 4.57],
    }
    manual_df = st.data_editor(
        pd.DataFrame(default_data),
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Stock":       st.column_config.TextColumn("Stock"),
            "Beta":        st.column_config.NumberColumn("Beta", format="%.4f"),
            "Weight (%)":  st.column_config.NumberColumn("Weight (%)", format="%.2f"),
        },
    )

    if st.button("Use this data"):
        manual_df = manual_df.dropna()
        manual_df["Beta"]      = pd.to_numeric(manual_df["Beta"],      errors="coerce")
        manual_df["Weight (%)"]= pd.to_numeric(manual_df["Weight (%)"],errors="coerce")
        manual_df = manual_df.dropna()
        df_input = manual_df.rename(columns={"Weight (%)": "Weight"}).copy()
        df_input["Weight"] = df_input["Weight"] / 100.0
        st.success(f"Using {len(df_input)} stocks from the table above.")

# ─────────────────────────────────────────────
# Optimization & Results
# ─────────────────────────────────────────────

def run_optimization(stocks, betas, current_weights, target, max_w):
    n = len(stocks)
    cap = max_w / 100.0

    def objective(w):
        return np.sum((w - current_weights) ** 2)

    def obj_grad(w):
        return 2 * (w - current_weights)

    constraints = [
        {"type": "eq", "fun": lambda w: np.dot(w, betas) - target},
        {"type": "eq", "fun": lambda w: np.sum(w) - 1.0},
    ]
    bounds = [(0.0, cap)] * n

    result = minimize(
        objective,
        x0=current_weights,
        method="SLSQP",
        jac=obj_grad,
        bounds=bounds,
        constraints=constraints,
        options={"ftol": 1e-10, "maxiter": 2000},
    )
    return result


def build_excel(df_result, old_beta, new_beta, target, inv):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False, sheet_name="Rebalancing Results")
        ws = writer.sheets["Rebalancing Results"]

        header_fill = PatternFill("solid", fgColor="1F3864")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_widths = [18, 10, 16, 22, 14, 22, 20, 22]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30

        total_row = ws.max_row
        total_fill = PatternFill("solid", fgColor="D6E4F0")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            if row[0].row % 2 == 0 and row[0].row != total_row:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F2F7FC")

        for cell in ws[total_row]:
            cell.fill = total_fill
            cell.font = Font(bold=True, name="Arial", size=10)

        ws.insert_rows(1)
        ws.insert_rows(1)
        ws["A1"] = f"Portfolio Rebalancing — Target Beta: {target}"
        ws["A1"].font = Font(bold=True, name="Arial", size=12, color="1F3864")
        ws["A2"] = (
            f"Current Beta: {old_beta:.4f}   →   Achieved Beta: {new_beta:.4f}"
            f"   |   Total Investment: {inv:,.0f} MAD"
        )
        ws["A2"].font = Font(italic=True, name="Arial", size=10, color="444444")

    buf.seek(0)
    return buf


if df_input is not None and len(df_input) >= 2:
    st.markdown("---")
    st.header("Optimization Results")

    stocks_list     = df_input["Stock"].tolist()
    betas_arr       = df_input["Beta"].to_numpy(dtype=float)
    weights_arr     = df_input["Weight"].to_numpy(dtype=float)

    # Warn if weights don't sum to ~1
    w_sum = weights_arr.sum()
    if not (0.98 <= w_sum <= 1.02):
        st.warning(
            f"Your weights sum to {w_sum:.2%} — they should sum to 100 %. "
            "The optimizer will still run but results may be unexpected."
        )

    result = run_optimization(stocks_list, betas_arr, weights_arr, target_beta, max_weight)

    if not result.success:
        st.error(f"Optimization warning: {result.message}")

    new_weights = result.x
    old_beta    = float(np.dot(weights_arr, betas_arr))
    achieved    = float(np.dot(new_weights, betas_arr))

    # ── Summary metrics ──────────────────────

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Current Beta",  f"{old_beta:.4f}")
    col2.metric("Target Beta",   f"{target_beta:.2f}")
    col3.metric("Achieved Beta", f"{achieved:.4f}", delta=f"{achieved - old_beta:+.4f}")
    col4.metric("Weights sum",   f"{new_weights.sum():.4f}")

    # ── Detailed table ────────────────────────

    changes     = new_weights - weights_arr
    df_out = pd.DataFrame({
        "Stock":            stocks_list,
        "Beta":             betas_arr,
        "Current Weight":   weights_arr,
        "New Weight":       new_weights,
        "Change":           changes,
        "Current Capital":  weights_arr * investment,
        "New Capital":      new_weights  * investment,
        "Capital Change":   changes      * investment,
    })

    st.dataframe(
        df_out.style
            .format({
                "Beta":            "{:.4f}",
                "Current Weight":  "{:.2%}",
                "New Weight":      "{:.2%}",
                "Change":          "{:+.2%}",
                "Current Capital": "{:,.0f} MAD",
                "New Capital":     "{:,.0f} MAD",
                "Capital Change":  "{:+,.0f} MAD",
            })
            .applymap(lambda v: "color: green" if isinstance(v, float) and v > 0
                      else ("color: red" if isinstance(v, float) and v < 0 else ""),
                      subset=["Change", "Capital Change"]),
        use_container_width=True,
        height=400,
    )

    # ── Charts ────────────────────────────────

    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            name="Current", x=stocks_list, y=weights_arr * 100,
            marker_color="#1F3864", opacity=0.8,
        ))
        fig_bar.add_trace(go.Bar(
            name="Rebalanced", x=stocks_list, y=new_weights * 100,
            marker_color="#2E75B6", opacity=0.9,
        ))
        fig_bar.update_layout(
            title="Weight Comparison (%)",
            barmode="group",
            yaxis_title="Weight (%)",
            legend=dict(orientation="h", y=-0.2),
            height=400,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    with chart_col2:
        fig_pie = go.Figure()
        fig_pie.add_trace(go.Pie(
            labels=stocks_list, values=new_weights * 100,
            name="Rebalanced weights",
            hole=0.35,
            textinfo="label+percent",
        ))
        fig_pie.update_layout(title="Rebalanced Portfolio Allocation", height=400)
        st.plotly_chart(fig_pie, use_container_width=True)

    # ── Export ────────────────────────────────

    # Build export DataFrame (formatted strings for Excel)
    df_export = pd.DataFrame({
        "Stock":                    stocks_list,
        "Beta (β)":                 betas_arr,
        "Current Weight":           [f"{v:.2%}" for v in weights_arr],
        "New Weight (Rebalanced)":  [f"{v:.2%}" for v in new_weights],
        "Weight Change":            [f"{v:+.2%}" for v in changes],
        "Current Capital (MAD)":    [f"{v:,.0f}" for v in weights_arr * investment],
        "New Capital (MAD)":        [f"{v:,.0f}" for v in new_weights  * investment],
        "Capital Change (MAD)":     [f"{v:+,.0f}" for v in changes     * investment],
    })
    summary_row = pd.DataFrame([{
        "Stock":                   "PORTFOLIO TOTAL",
        "Beta (β)":                f"{achieved:.4f}",
        "Current Weight":          f"{weights_arr.sum():.2%}",
        "New Weight (Rebalanced)": f"{new_weights.sum():.2%}",
        "Weight Change":           "—",
        "Current Capital (MAD)":   f"{investment:,.0f}",
        "New Capital (MAD)":       f"{new_weights.sum() * investment:,.0f}",
        "Capital Change (MAD)":    "—",
    }])
    df_export = pd.concat([df_export, summary_row], ignore_index=True)

    excel_buf = build_excel(df_export, old_beta, achieved, target_beta, investment)

    st.download_button(
        label="Download rebalanced portfolio (.xlsx)",
        data=excel_buf,
        file_name="rebalanced_portfolio.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info(
        "Upload an Excel file in the first tab **or** fill in the table in the second tab "
        "and click **Use this data** to see the rebalancing results."
    )
